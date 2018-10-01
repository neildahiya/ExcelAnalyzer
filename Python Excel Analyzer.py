from xlrd import open_workbook
import xlwt
#Finding location of row column index of reference sheet
import openpyxl
path = "C:\\Users\\Neil Dahiya\\Documents\\xl3.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
i = 1
j = 1
for i in range(1,6) :
    cell_obj = sheet_obj.cell(row=i, column=j)
    print(cell_obj.value)
    E = cell_obj.value
    
    book = open_workbook("xl1.xlsx")
    
#Looping through xl1 list
    for sheet in book.sheets():
        for rowidx in range(sheet.nrows):
            row = sheet.row(rowidx)
            for colidx, cell in enumerate(row):
                if cell.value == E :
#                    print (sheet.name)
                    
                    print (rowidx+1)
                    print (colidx+1)
                    
                    cell2 = book.sheet_by_index(0).cell(rowidx,colidx+1).value
#                    cell3 = book.sheet_by_index(0).cell(rowidx,colidx+2).value
                    
                    print (cell2)
#                    print (cell3)
                    
#                    a = colidx+2
#                    b = rowidx+1
#                    c = colidx+3
#                    d = rowidx+1
#                    
#                    
                    sheet_obj.cell(row=i, column=j+1).value = cell2
#                    sheet_obj.cell(row=i, column=j+2).value = cell3
#                    print (sheet_obj.cell(row=1, column=1).value)
                    wb_obj.save("Output.xlsx")
                else: continue           













